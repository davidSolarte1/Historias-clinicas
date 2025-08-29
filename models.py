from database import get_connection, hash_password
import pandas as pd
from email_utils import enviar_correo
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---- Usuarios ----
def verificar_usuario(email, password):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, nombre, rol FROM usuarios WHERE email=%s AND password_hash=%s", (email, hash_password(password))
        )
        user = cursor.fetchone()
        conn.close()
        return user  # (id, nombre, rol) o None
    except Exception as e:
        print("❌ Error al conectar con MySQL:", e)

def registrar_usuario(cedula, nombre, email, password, rol):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
        INSERT INTO usuarios (cedula, nombre, email, password_hash, rol)
        VALUES (%s, %s, %s, %s, %s)
        """, (cedula, nombre, email, hash_password(password), rol))
        conn.commit()
        mensaje = (
            f"Cordial saludo,\n\n"
            f"{nombre},\n\n"
            f"Se ha creado tu cuenta en el sistema SIREH.\n"
            f"Sus credenciales son:\n" 
            f"📧 Usuario: {email}\n"
            f"🔑 Contraseña: {password}\n\n"
            f"Atte,\n\n"
            f"____________________________\n"
            f"ADMINISTRADOR ARCHIVO CLÍNICO\n"
            f"Correo-E: archivoclinico.hc@hudn.gov.co\n"
        )
        enviar_correo(
            email,
            "Credenciales de acceso - SIREH",
            mensaje
        )
        return True
        
    except:
        return False
    finally:
        conn.close()

def usuario_existe(cedula: str) -> bool:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM usuarios WHERE cedula = %s LIMIT 1", (cedula.strip(),))
    existe = cur.fetchone() is not None
    conn.close()
    return existe

def obtener_usuario_por_cedula(cedula: str):
    """
    Devuelve (id, nombre, email, rol) del usuario con ese email (case-insensitive),
    o None si no existe.
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, cedula, nombre, email, rol
        FROM usuarios
        WHERE cedula = %s
        LIMIT 1
    """, (cedula.strip(),))
    row = cur.fetchone()
    conn.close()
    return row  # None o (id, nombre, email, rol)


def actualizar_usuario(cedula: int, nombre: str, email: str, password: str or None, rol: str) -> bool:
    """
    Actualiza nombre, email, rol y opcionalmente la contraseña.
    Si password es None o "", NO cambia la contraseña.
    Devuelve True si actualizó, False en error.
    """
    conn = get_connection()
    cur = conn.cursor()
    try:
        if password:  # cambiar contraseña
            cur.execute("""
                UPDATE usuarios
                SET nombre = %s, email = %s, password_hash = %s, rol = %s
                WHERE cedula = %s
            """, (nombre.strip(), email.strip().lower(), hash_password(password), rol, cedula))
        else:  # sin cambiar contraseña
            cur.execute("""
                UPDATE usuarios
                SET nombre = %s, email = %s, rol = %s
                WHERE cedula = %s
            """, (nombre.strip(), email.strip().lower(), rol, cedula))
        conn.commit()
        return cur.rowcount > 0
    except Exception as e:
        print("Error actualizar_usuario:", e)
        return False
    finally:
        conn.close()

def obtener_enfermeros():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre FROM usuarios WHERE rol='user'")
    data = cursor.fetchall()
    conn.close()
    return data

# def obtener_historias(fecha_ini, fecha_fin, enfermero_id=None):
#     conn = get_connection()
#     cursor = conn.cursor()

#     if enfermero_id and enfermero_id != None:
#         cursor.execute("""
#             SELECT h.id, h.numero_carpeta, h.cedula_paciente, h.nombre_paciente,  h.apellido_paciente, h.servicio, h.fecha_recepcion, u.nombre, h.estado, h.fecha_devolucion, h.fecha_entrega_nueva
#             FROM historias_clinicas h
#             JOIN usuarios u ON h.usuario_registro = u.id
#             WHERE h.fecha_recepcion BETWEEN %s AND %s AND u.id = %s
#         """, (fecha_ini, fecha_fin, enfermero_id))
#     else:
#         cursor.execute("""
#             SELECT h.id, h.numero_carpeta, h.cedula_paciente, h.nombre_paciente,  h.apellido_paciente, h.servicio, h.fecha_recepcion, u.nombre, h.estado, h.fecha_devolucion, h.fecha_entrega_nueva
#             FROM historias_clinicas h
#             JOIN usuarios u ON h.usuario_registro = u.id
#             WHERE h.fecha_recepcion BETWEEN %s AND %s
#         """, (fecha_ini, fecha_fin))

#     data = cursor.fetchall()
#     conn.close()
#     return data

def obtener_servicios():
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT DISTINCT servicio FROM historias_clinicas ORDER BY servicio ASC")
        return [row[0] for row in cur.fetchall()]
    finally:
        cur.close()
        conn.close()

def obtener_historias(fecha_ini, fecha_fin, enfermero_id=None, servicio=None, estado=None):

    conn = get_connection()
    cur = conn.cursor()
    try:
        sql = """
        SELECT
            h.id,
            h.numero_carpeta,
            h.cedula_paciente,
            h.nombre_paciente,
            h.apellido_paciente,
            h.servicio,
            h.fecha_recepcion,
            u.nombre AS usuario,
            h.estado, 
            h.fecha_devolucion, 
            h.fecha_entrega_nueva
        FROM historias_clinicas h
        INNER JOIN usuarios u ON u.id = h.usuario_registro
        WHERE
          (
            (h.fecha_recepcion BETWEEN %s AND %s)
            OR (h.fecha_devolucion BETWEEN %s AND %s)
            OR (h.fecha_entrega_nueva BETWEEN %s AND %s)
          )
        """
        params = [fecha_ini, fecha_fin, fecha_ini, fecha_fin, fecha_ini, fecha_fin]

        if enfermero_id:  # si viene None, no aplica
            sql += " AND u.id = %s"
            params.append(enfermero_id)

        if servicio and servicio != "Todos":
            sql += " AND h.servicio = %s"
            params.append(servicio)

        if estado and estado != "Todos":
            sql += " AND h.estado = %s"
            params.append(estado)

        sql += " ORDER BY COALESCE( h.fecha_recepcion) DESC"

        cur.execute(sql, tuple(params))
        return cur.fetchall()
    finally:
        cur.close()
        conn.close()

def agregar_observacion(historia_id, texto):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE historias_clinicas
        SET observacion=%s, estado='devuelta', fecha_devolucion=%s, fecha_entrega_nueva = '-'
        WHERE id=%s
    """, (texto, date.today().isoformat(), historia_id))
    conn.commit()

    # Obtener email
    cursor.execute("""
        SELECT u.email, h.numero_carpeta,  h.observacion
        FROM historias_clinicas h
        INNER JOIN usuarios u ON h.usuario_registro = u.id
        WHERE h.id=%s
    """, (historia_id,))
    resultado = cursor.fetchone()
    conn.close()

    if resultado:
        correo_enfermero, numero_carpeta, observacion= resultado
        mensaje = (
            f"Cordial saludo,\n\n"
            f"Se ha agregado una observación a la historia clínica:\n"
            f"📂 Número de carpeta: {numero_carpeta}\n"
            f"📝 Observación: {observacion}\n\n"
            "Por favor, revisa el sistema para realizar las correcciones necesarias.\n\n"
            f"Atte,\n\n"
            f"____________________________\n"
            f"ADMINISTRADOR ARCHIVO CLÍNICO\n"
            f"Correo-E: archivoclinico.hc@hudn.gov.co\n"


        )
        enviar_correo(
            correo_enfermero,
            "Observación en historia clínica",
            mensaje
        )

def exportar_reporte_excel(ruta_salida="reporte.xlsx"):
    conn = get_connection()
    query = """
    SELECT h.id, h.numero_carpeta, h.cedula_paciente, h.nombre_paciente,  h.apellido_paciente, h.servicio, h.fecha_recepcion, 
           u.nombre as enfermero, h.observacion, h.estado, h.fecha_devolucion, h.fecha_entrega_nueva
    FROM historias_clinicas h
    INNER JOIN usuarios u ON h.usuario_registro = u.id
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    if df.empty:
        return False

    df.to_excel(ruta_salida, index=False, sheet_name="Historias")

    # openpyxl para darle estilo
    libro = openpyxl.load_workbook(ruta_salida)
    hoja  = libro.active

    # --- Estilo encabezado ---
    encabezado_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for cell in hoja[1]:
        cell.font = encabezado_font
        cell.fill = fill
        cell.alignment = center

    # --- Ajuste ancho automatico ---
    for column_cells in hoja.columns:
        max_length = 0
        col = column_cells[0].column_letter  
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        hoja.column_dimensions[col].width = max_length + 2

    libro.save(ruta_salida)
    return True

def marcar_entregada(historia_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE historias_clinicas
        SET estado='entregada'
        WHERE id=%s
    """, (historia_id,))
    conn.commit()
    conn.close()


# Funciones para el usuario
# ---- Historias ----
def guardar_historia(numero_carpeta, cedula, nombre, apellido, servicio, usuario_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO historias_clinicas (numero_carpeta, cedula_paciente, nombre_paciente, apellido_paciente, servicio, fecha_recepcion, usuario_registro)
    VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (numero_carpeta, cedula, nombre, apellido,  servicio,date.today().isoformat(), usuario_id))
    conn.commit()
    conn.close()

def obtener_historias_devueltas(usuario_id: int):
    """Historias del usuario en estado 'devuelta'."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, numero_carpeta, cedula_paciente, nombre_paciente, apellido_paciente, servicio, fecha_recepcion, observacion
        FROM historias_clinicas
        WHERE usuario_registro = %s AND estado = 'devuelta'
        ORDER BY date(fecha_recepcion) DESC
    """, (usuario_id,))
    rows = cur.fetchall()
    conn.close()
    return rows

def obtener_historia_por_id(historia_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, numero_carpeta, cedula_paciente, nombre_paciente, apellido_paciente, servicio, fecha_recepcion, observacion, estado
        FROM historias_clinicas
        WHERE id = %s
    """, (historia_id,))
    row = cur.fetchone()
    conn.close()
    return row

def actualizar_historia_revisada(historia_id: int, numero_carpeta: str, cedula_paciente: str,nombre: str, apellido: str, servicio: str):
    """
    Actualiza la historia devuelta; por defecto pone la fecha de hoy
    y cambia estado a 'revisada'.
    """

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE historias_clinicas
        SET numero_carpeta = %s, cedula_paciente = %s, nombre_paciente = %s, apellido_paciente = %s, servicio = %s, estado = 'revisada', fecha_entrega_nueva=%s
        WHERE id = %s
    """, (numero_carpeta, cedula_paciente, nombre, apellido, servicio,  date.today().isoformat(), historia_id))
    conn.commit()
    conn.close()